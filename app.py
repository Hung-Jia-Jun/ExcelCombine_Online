from flask import Flask,request
from flask import render_template
import time
import sys
import json
from flask_cors import CORS
from flask_socketio import SocketIO, emit
from os.path import dirname, abspath 
from openpyxl import load_workbook,Workbook
import os
from os import listdir
app = Flask(__name__)
CORS(app)
socketio = SocketIO(app)

# class sensor_ph(db.Model):
# 	DateTime = db.Column(db.String(255), primary_key=True)
# 	Data = db.Column(db.String(255))
  
# 	def __init__(self, DateTime, Data):
# 		self.DateTime = DateTime
# 		self.Data = Data
#------------------------------------------------------------------------------------------------------

@app.route("/index")
def index():
	return render_template('index.html')

#取得當前指令的列表
# @app.route("/queryCommandList")
# def queryCommandList():
# 	#依照ID排序
# 	scheduleLi = schedule.query.order_by(schedule.id.asc()).all()
# 	Command_li = []
# 	for ele in scheduleLi:
# 		if ele != None:
# 			scheduleCommand = {
# 				'id' : ele.id,
# 				'PositionX' : ele.PositionX,
# 				'PositionY' : ele.PositionY,
# 			}
# 			Command_li.append(json.dumps(scheduleCommand))
# 	return json.dumps(Command_li)


@socketio.on('Upload_Stream')
def Upload_Stream(msg):
	filename = msg["filename"]
	binary = msg["data"]
	f = open(filename, "wb")
	f.write(binary)
	f.close()
	# socketio.emit('ImageStream', {'data': str(encoded_string, encoding = "utf-8")})


@socketio.on('Start_Combine')
def Start_Combine(msg):
	command = msg["data"]
	if command == "Start Combine":
		MergeRows = []
		title = []
		for filename in listdir(os.path.dirname(os.path.abspath(__file__))):
			if "xlsx" in filename:
				workbook = load_workbook(filename)
				sheets = workbook.get_sheet_names()         
				booksheet = workbook.get_sheet_by_name(sheets[0])
				
				rows = booksheet.rows
				columns = booksheet.columns
				index = 0

				title_List=[]
				title = list(booksheet.rows)[0]
				for i in range(len(title)):
					title_List.append(title[i].value)
				title = title_List
				for row in rows:
					colLine = []
					if index > 0:
						for col in row:
							colLine.append(str(col.value))
						MergeRows.append(colLine)
					index += 1
			
		book = Workbook()
		sheet = book.active

		sheet.append(title)
		for row in MergeRows:
			sheet.append(row)

		book.save(os.path.dirname(os.path.abspath(__file__)) + '/static/combine.xlsx')
		socketio.emit('Combine_Response', {'data': "OK"})


@socketio.on('Clear_History')
def Clear_History(msg):
	command = msg["data"]
	if command == "clear":
		for filename in listdir(os.path.dirname(os.path.abspath(__file__))):
			if "xlsx" in filename:
				os.remove(filename)
		try:
			os.remove(os.path.dirname(os.path.abspath(__file__)) + '/static/combine.xlsx')
		except:
			pass
if __name__ == "__main__":
	app.run(host='0.0.0.0',port='8000')
	socketio.run(app,debug=True)
